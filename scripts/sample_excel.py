import openpyxl as xl
from openpyxl.chart import BarChart, Reference

is_file = False


def process_workbook():
    while True:
        try:
            pick_a_file = input("Enter a xlsx file>>")
            wb = xl.load_workbook(pick_a_file)
            sheet = wb['Sheet1']
            global is_file
            is_file = True
        except:
            print("No such file found!")
        if is_file:

            for satir in range(2, sheet.max_row + 1):
                cell = sheet.cell(satir, 3)
                new_price = cell.value * 0.9
                new_cell = sheet.cell(satir, 4)
                new_cell.value = new_price

            values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4,
                               max_col=4)  # dosyamdaki 2/3/4.. satırların 4. sütunları
            chart = BarChart()  # bar chart class oluşturdum.
            chart.add_data(values)
            sheet.add_chart(chart, 'e2')
            choice = input("Press y to overwrite or type anyname to create a copy>")
            if choice.lower() == "y":
                wb.save(pick_a_file)
                print("Saved.")
                break
            else:
                wb.save(choice + ".xlsx")
                print(f"Saved as {choice}.xlsx")
                break


process_workbook()
