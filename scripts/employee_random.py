#rastgele işçi üretip excele yazdırır.
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Font,Color,colors,Alignment
from random import randint,choice
class Employee:
    def __init__(self,name,last,pay=30000):
        self.name = name.title()
        self.last = last.title()
        self.pay = pay
        self.email = name.title()+'_'+last.title()+'@company.com'
        self.pay = randint(24000,40000)
        self.applied = False
    def describe(self):
        print(f'Name: ',self.name,'\nSurname:',self.last,'\nE-mail:',self.email,'\nPay:',self.pay)

    @classmethod
    def userstr(cls,str):
        name,last,pay = str.split('-')
        pay = int(pay)
        return cls(name,last,pay)
    def apply_raise(self):
        self.pay *=0.4
        self.pay = int(self.pay)
        self.applied = True

hasan = Employee('hasan','aslan',30000)
mehmet = 'Mehmet-aslan-30000'
mehmet1 = Employee.userstr(mehmet)

list = [mehmet1,hasan]
names = ['Yakup','Enes','Murat','Seda','Tuğçe','Sakine','Samet','Sefer','Yunus','Melike','Mustafa','Ayşe']
used_names = []
surnames = ['Aslan','Gürle','Tuğra','Cingöz','Reis','Kayaoğlu','Kaplan','Eril','Urgan','Kayaoğlan','Özerler']
while len(list)!=50:
    chance_to_apply = randint(1,3)
    fullname = ''
    name = choice(names)
    if name in used_names:
        continue
    surname = choice(surnames)
    pay = randint(24000,40000)
    full = name+'-'+surname+'-'+str(pay)
    random_user=Employee.userstr(full)
    list.append(random_user)
    if chance_to_apply==2:
        random_user.apply_raise()
try:
    wb = xl.load_workbook('users.xlsx')
    sheet = wb['Sheet1']
except:
    wb = Workbook()
    worksheet = wb.active
    worksheet.title = 'Sheet1'
    sheet = wb['Sheet1']
cell = sheet.cell(1,1)
cell.font = Font(color=colors.BLUE, italic=False, bold=True)
cell.value = 'Name'
cell = sheet.cell(1,2)
cell.font = Font(color=colors.BLUE, italic=False, bold=True)
cell.value = 'Surname'
cell = sheet.cell(1,3)
cell.font = Font(color=colors.BLUE, italic=False, bold=True)
sheet.column_dimensions['C'].width = 30
cell.value = 'E-mail'
cell = sheet.cell(1,4)
cell.font = Font(color=colors.BLUE, italic=False, bold=True)
cell.value = 'Pay'
cell = sheet.cell(1,5)
cell.value = 'Applied For Raise'
cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
b = 2
for a in list:
    cell = sheet.cell(b,1)
    cell.value = a.name
    cell = sheet.cell(b,2)
    cell.value = a.last
    cell = sheet.cell(b,3)
    cell.value = a.email
    cell = sheet.cell(b,4)
    cell.value = str(a.pay)+'$'
    cell = sheet.cell(b,5)
    if a.applied == False:
        cell.value = 'No'
    else:
        cell.value = 'Yes'
    b+=1

wb.save('users.xlsx')
input('Completed')