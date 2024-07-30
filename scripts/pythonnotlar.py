^^^^Random^^^^
import random
sayi=random.randint(a,b) # a' b sayısından küçük olmalı.
dizi=[1,2,3,4,5]
a=random.choice(dizi)
    
    ARRAYS

dizi=["a","b","c","d","e","f"]
print(dizi[3:]) #### 3. index dahil, 3'ten sonrasını yazar..
print(dizi[2:3]) ##### 2 dahil 3 hariç.
string= "python"
print(string[3:]) #### "ing"
len(string)########>>>>>> karakter sayısı.==6
if len(input)!=1: # eğer girdi 1 haneli değilse..
print(*dizi) # yan yana noktasız virgülsüz yazar.

print(a, end=",")#### While komutunda kullanabilirim. Loop sırasında alt alta yazmak yerine virgülle ayırır.
dizi=["Hasan","Yakup","Seda","Mustafa","Hasan"]

^^Listedeki aynı ögeleri siler.
dizi = list(dict.fromkeys(dizi))


print(*dizi)# parantez olmadan yazdırır. print(*dizi, sep="-") bu da "-" ile elementleri ayırır
print(dizi.index("hasan")) ## hangi indexte olduğunu gösterir.
dizi.append("haha") ## dizi dizisinin sonuna 'haha' karakterini ekler.
dizi.insert(1,"Mehmet") ##1 nolu dizeye ''Mehmet'' ekler.
dizi.extend(başkabirdizi) ## iki  diziyi birleştirir.
dizi.remove('mehmet')##Mehmet'i siler
dizi.pop()## sondakini siler
dizi.sort() ___ print(dizi)#    alfabetik olarak sıralar.
dizi.reverse() ## elementleri tersten başa sayar. En sondaki 0
print(dizi.count("Hasan"))## dizide kaç adet bu elementten olduğunu gösterir.
dizi2 =dizi.copy() ## diziyi kopyalar.
a=input("Adın ne?").strip() #boşlukları kaldırır.
^^^^^String format^^^
print(f"haha {variable}")
print("haah {0}".format(variable))

^^^^^iki liste arasındaki farkı bulma^^^^
liste1= [1,2,3,4,5,6,7]
liste2= [2,3,4,1,5,2,5,6,7,8,2]

setliste1 = set(liste1)
setliste2=set(liste2)
sonuc = setliste2.difference(setliste1)
print(f"Liste1: {liste1}\nListe2: {liste2}")
print(f"Elements that in liste2 but not in liste1: {sonuc}")
#>Output = 8

^^^^^^^^^^
name = 'joHN smİth'
print(name.title()) # John Smith yazdıracak 
^^^^^^^^^^^^^^^^^^^^^ Bir stringe harfleri yazıp onları karışık bir şekilde aldım^^^
import random
list = []
keys = "ABCDEFGHIJKLMNOPRSTUVYZQWX1234567890"
for a in keys:
   list.append(a)
random.shuffle(list)
print(*list,  sep="")



^^^^^^^^^^^^DEĞİŞKENLER- Variables^^^^^^^
global, local, nonlocal
# fonksiyon içinde bir değişken oluşturursam bu local olur. Fonksiyon bitince onu bulamam. Bunun için fonksiyon tanımlanmadan önce global yapmalıyım;
global x
x=int(input("Bir sayı gir: "))

fonksiyon içi bir değişkeni, fonskiyon dışına çıkarmak istersem global olmalı
eğer içiçe bir fonksiyon yazdıysam ve içteki fonksiyonun değişkeninin ana fonksiyonda da istersem nonlocal kullanırım.
def fonk():
	def inner():
		nonlocal x
		x=2
	print(x)
^^^^^^^^^^^^^^^^^^^input ile ilgili >> print(input("enter a word with a space").replace(" ","")) ## Boşluğu siler. Ya da istedğimi yazar.
 



^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^While^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
a=2
while a>1:
    statements.. # a 1 sayısından büyük, True. O zaman alttaki kodlarını uygular. Break yerine continue yazarsan kod yukardan tekrar başlar.
    continue


^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^TUPLES^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^

coordinates(4,5) ## Tuple içindeki veri değiştirilemez. Parantezle gösterilir. Listelerden daha hızlıdır. 
coordinates[(2,3),(4,5),(6,7)] # bu şekilde de yapabiliriz.
tekElemanlıTuple = (1,)


^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^SETS^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
#Unique. Indexleri yoktur ve karmasık sıradadırlar. Hızlıdırlar.
mySet = {"Hasan","Mehmet","Seda"}
mySet.add("Ceulan")
mySet.update(['a','b','c'])
mySet.discard('c')
mySet.pop() #son elemanı siler.
mySet.clear() #temizler.
print(mySet)

^^Union

setA = {1,2,3,4,5}
setB = {1,3,4,6,7,8}

print(setA | setB)#duplicate verileri atar.
print(setA.union(setB)) #Ya da böyle yapabilirim.
print(setB.union(setA))

^^intersection #ortak verileri bulmak

print(setA & setB)
print(setA.intersection(setB))
print(setB.intersection(setA))

^^Difference
print(setA - setB)
print(setA.difference(setB)) #setA da olup diğerinde olmayan elemanlar.
print(setB.difference(setA))
 #direkt ikisinin farklarını verir.
^^Symmetric Difference
print(setA ^ setB)
print(setA.symmetric_difference(setB))
print(setB.symmetric_difference(setA))


^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^FUNCTIONS^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
def say_hello:
	print("hello user") ### fonksiyon oluştu.
say_hello() ## Bu da fonksiyonu çağırır ve işler. 
def say_hello(name, age):
    print("Hello "+name+". You are "+age+" years old.")
say_hello("Hasan","21") ##### parantez içindeki name=parameter. Output>>>>> "Hello Hasan. You are 21 years old.""
## fonksiyonu çağırırken sayıyı string yerine int olarak vermek istersek fonksiyon ayarında age kısmını str(age) olarak değiştir. Sayıyı int verir string alırsın.
^^^			*args
not = "Fonksiyonu tanımlarken kullandığım değişken parameterdir. Kullanırken ise verdğim değer argumentdir."
not 2= 'Eğer fonksiyona kaç argument verileceğini bilmiyorsam, *args kullanırım. Bu bir tuple dizisi gibi olur.'

def myfunc(*kids):
	print("The youngest child is",kids[2])

myfunc("Hasan","Hüseyin","Mustafa")

^^Bir fonksiyon parameteri için şunları da kullanabilirim:

def album(artist,title,number_of_tracks = None): # Number_of_tracks böyle yazdıö çünkü belki o değeri her zaman eklemem.
    dic= {'Artist':artist.title(),'Album title':title.title()}
    if num:
        dic['Number of Tracks'] = number_of_tracks
    return dic

hasan = album('hasan aslan','gece yolcuları',number_of_tracks=5)



^^^		**kwargs
not =' *args için tuple demiştik. Bunda ise Dictionary alır.'
	
def myfunc(**user_info):
	print("Your last name is",user_info["last_name"])

myfunc(age = 21, last_name ="Aslan",name= "Hasan")

^^^^^^^^^^^^^^fonksiyon yinelemeri / Recursion #Altlarda başka bir yerde örneğim var. 
def recur(x):
    if x>0:
        result = x+ recur(x-1)
        print(result)
    else:
        result = 0
    return result


recur(10)



^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^TIME - DATETIME-------------------------------

def format_time():
    import datetime
    x = datetime.datetime.now()
    output = ''
    convert_str = x.strftime('%x')
    output+= convert_str[3:5]+'/'
    output+=convert_str[0:2]+'/'
    output+=convert_str[6:8]
    clock_time = str(x.strftime('%X'))
    output +=' | '+clock_time
    print(output)
#kendi hazırladığım zaman şeysi.
^^^^^^^^^Geçen Süreyi hesaplama
import time
start = time.time()
some code..


end = time.time()
elapsed_time = end-start
elapsed_time = round(elapsed_time,2)

^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^RETURN STATEMENT^^^^^^^^^^^^^^^^^^^^^^^^^^^
def cube(num):
	return num*num*num
print(cube(3)) ### ouput>>27. return koymasaydım output=none olacaktı.Parametere değer vermediğim için öyle olur. Return, parameteri her gördüğünde verilen değere bakıp ona değer verir.
-----------------
def max_num(num1,num2,num3):
    if num1>=num2 and num1>=num3:
        return num1
    elif num2>=num1 and num2>=num3:
        return num2
    else:
        return num3
print(max_num(3,40,5)) ## maksimum numarayı bulur.



^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^IF STATEMENT^^^^^^^^^^^^^^^^^^^^^^^^^^^^^  
is_male= True ## Boolean
is_tall=True
if is_male or is_tall: ### 'or' yerine 'and' kullanırsam iki şartın da olması gerekir. 'or' için birinin olması yeter.
	print("You are a male or tall or both..")
elif is_male and not(is_tall):
	print("You are a short male.") ## erkek ve uzun değil ise bunu yazar.
else:
	print("You are not a male or short or both.")########## You are a male or tall or both..
^^^^ if statement verirken else komutunun altına else durumunda olacak conditionları mutlaka gir.
if input=="male" or input=="Male": ## 2 opsiyondan biri
if input.lower()="male" ## Girdi büyüklü küçüklü olsa dahi kabul eder.
ya da if input in ['male', 'Male']: ## ama en iyisi lower yapmak.


number = 2
mes = "number is five" if number==5 else ("no")
print(mes)


^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^DICTIONARIES^^^^^^^^^^^^^^^^^^^^^^^^^^3
aylar={
    "Jan":"January",
    "Feb":"February",

}
print(aylar["Jan"]) # OUTPUT>>>> January
ya da print(aylar.get("Jan")) ## olarak da yapabiliriz.
print(aylar.get("Mar","Not such a key"))### Mar keyini bulamadığı için Not such a key diyecek.

aylar["mar"] = "March"#Yeni bir değer eklemek için

del aylar["mar"] ## Siler.
^^^Ya da
aylar.pop("mar")

aylar.update({"Jan":"İlk ay"}) ## Bu da mevcut değerleri değiştirir.
march = aylar.pop("mar")
print(march) # değeri yazdırır. 

del aylar # sözlüğü siler.
aylar.clear()##her şeyi sıfırlar.


print(aylar.items())
^^Dictionary loop
for key,value in aylar.items():
	print(key,value)



^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^For Loop^^^^^^^^^^^^^^^^^^^^
##Atadığın değişkenin her bir karakterini değişken olarak tanımlar
for letter in "hasan":
	print(letter)### output alt alta h a s a n olur.
	#################3
friends=["hasan","batu","ebrkay","karhan"]
for friend in friends:
    print(friend) ### Alt alta liste ypapar
###############
friends=["hasan","batu","berkay","karhan"]
for friend in range(len(friends)):
    print(friends[friend]) ######## Eğer print(friends) dersem alt alta 0-1-2-3 yazar. Ama böyle yapınca liste yapar
    ^^^^Normal olarak for döngüsünü sayılar için kullanamayız ama string yardımı alabiliriz:
 sayılar = "123456789"
for sayı in sayılar:
    print(int(sayı) * 2)### her sayıyı alt alta 2 ile çarparak yazar.
if not any(c in digits for c in input): ## loop olmadan işe yarar. Password validatorumda görebilirim. Bunun için from string import digits, punctuation



for a in range(2,30,3):
	print(a) ## 2-29 arası 3erli atlar.

^^^^^^^^^^
def raise_to_power(base_num,pow_num):
    result=1
    for i in range(pow_num):
        result=result*base_num
    return result
print(raise_to_power(2,3)) ## output 8 . Sayıyı 2'nin 3ün gücüne alır.

^^^^^^^^^^^^^^^^Sırayla animasyon gibi yazdırma^^^^^^^^^^^33

def text(text):
    import sys
    import time
    for a in text:
        print(a,end='')
        sys.stdout.flush()
        time.sleep(0.1)





^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^2D lISTS^^^^^^^^^^^^^^^^^^
number_grid=[                 #### liste içi listeler. ilk numara liste indexi, ikinci ise liste ici
    [1,2,3],
    [4,5,6],
    [7,8,9],
    [0]
]
print(number_grid[1][1])

^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^MATRİX SUM^^^^^^^^^^^^^
x = [[1,3,5],[2,4,1],[1,5,7]]

y = [[3,3,4],[2,4,1],[3,5,4]]

sonuc = [[0,0,0],[0,0,0],[0,0,0]] #ikisinin toplamı

for i in range(len(x)):
    for a in range(len(y)):
        sonuc[i][a] = x[i][a]+y[i][a]


^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^List/tuple slices^^^^^^^^^^^^^
squares = [0, 1, 4, 9, 16, 25, 36, 49, 64, 81] # sayıların kareleri
print(squares[0:3]) ## 0. indexten 3. indexe kadar alır. 3. index dahil değildir. (aradaki fark kadar element gibi düşün.)
print(squares[:3]) ## aynı mantık 
print(squares[3:]) # 3. index dahil, sonrasını yazdırır
print(squares[::2]) ## 2'şerli olarak atlar. 0,4,16,36,64
print(squares[2:8:3]) # 2. elementten 8. elemente kadar 3^erli gider. 8 dahil değildir.



^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^List comprehension^^^^^^^^^^^
cubes=(i**3 for i in range(5))
print(cubes) # direkt 5 e kadar olan sayıların küpünü verir.
# if komutu da uygulanabilir.
evens=[i**2 for i in range(10) if i**2 % 2 == 0] # 0-10 arasındaki her sayı için şu işlemi yapar: eğer sayının kendisiyle çarpımı 2'ye bölününce kalan 0 ise, o sayının kendisiyle çarpımını yazar.
print(evens)





^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^String Formatting^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
nums = [4, 5, 6]
msg = "Numbers: {0} {1} {2}". format(nums[0], nums[1], nums[2])
print(msg) 
a = "{x}, {y}".format(x=5, y=12)

print(", ".join(["spam", "eggs", "ham"]))#prints "spam, eggs, ham"
print("Hello ME".replace("ME", "world"))#prints "Hello world"
print("This is a sentence.".startswith("This"))# prints "True"
print("This is a sentence.".endswith("sentence."))# prints "True"
print("This is a sentence.".upper())# prints "THIS IS A SENTENCE."
print("AN ALL CAPS SENTENCE".lower())#prints "an all caps sentence"
print("spam, eggs, ham".split("a")) ## a harflerini siler.
msg="merhaba dünya".split() # aynı zamanda ayrı yazılan her bir kelimeyi bir element yapar. Dizi gibi

print(min(1, 2, 3, 4, 0, 2, 1))# 0
print(max([1, 4, 9, 2, 5, 6, 8])) #9
print(abs(-99)) #99 absolute değeri alır.
print(abs(42)) #42
print(sum([1, 2, 3, 4, 5])) #15 toplam.
a=min([sum([11, 22]), max(abs(-30), 2)])# 11+22=33, diğer tarafta 30>2. Elimizde sırayla 33 ve 30 var. Dikkat etmemiz gereken nokta, solda en DIŞARIDA
#min komutu var. Dolayısıyla en küçük sayı olan 30'u almak durumundayım.



^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^List Functions^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
nums = [55, 44, 33, 22, 11]

if all([i > 5 for i in nums]):
   print("All larger than 5")

if any([i % 2 == 0 for i in nums]):
   print("At least one is even")

for v in enumerate(nums):
   print(v) # alt alta indexi ve elementini yazdırır.   



^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^Basic Translator^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
def translate(phrase):
    translation=""
    for letter in phrase:
        if letter in "AEIOUaeiou":
            translation=translation+"g"
        else:
            translation=translation+letter
    return translation
print(translate(input("Enter: ")))  #### yazdığım her ünlü harfi 'g' harfi ie değiştirir.



^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^Error Handling^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
try:
num1=int(input("Enter a num:"))
print(num1)
except:
    print("Enter a number, not a character")  ##Normalde sayı girmezsem program çöker. Ama burada hata olunca except kod blogunun altındaki kodlar çalışacak
 ^^Hata türleri; ImportError, IndexError, NameError, SyntaxError, TypeError, ValueError
 try:
    a=10/0
    print(a)
except:
    print("Sıfıra bölemeyeceğim için bu mesajı verecek. Bu kısım olmazsa program çöker")   ## except ZerodivisionError şeklinde de olabilir.
finally:
    print("This code will run whatever happens.")

^^^^
try:
	a =int(input('enter a number: '))
except:
	print('Enter a number!')
else:
	print('If try code works, this will be printed.')



    ^^^^^^^^^^^^^^^^^^^^^Raising errors^^^^^^^^^^^
    a=1
    raise ValueError("Bu mesajı yayınlayarak bir ValueError oluşturur")
    ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^



    ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^Files^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
Myfile= open("dosyaadi.txt", "r")# r= read, w=write, a=append, r+=read and write. Değişken atanabilir. 
print(Myfile.readable()) # Dosyanın okunup okunamadığına dair boolean değeri verir.
print(Myfile.read())  ###dosyanın içindekileri yazdırır.
print(Myfile.readline()) #ilk sırayı yazdırır ikincisine geçer. Kodu tekrar yazarsan 2yi yazdırıp 3e geçer.
print(Myfile.readlines()) # tüm satırları bir array içinde yazar.
Myfile.close()
    ^^^^
my_data= open("example.txt", "r")
for a in my_data.readlines():
    print(a)
my_data.close() # Bu da basitçe sıraları yazdırır.
^^^^^^^^^^^^^^^^^^^^^^^^ # alttaki komutlar kullanıcıya bir dosya açtırtır ve okutur.^^^^^^^
filename = input("Enter a filename: ")
with open(filename) as f:
   text = f.read()
print(text)


^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^Writing to files^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
myfile=open("example.txt","a")
myfile.write("\nBu yazılan en sona yazılacak")
myfile.close()
^^^^^^^^^^^^^^^^^^^^^^^^^3
myfile=open("example.txt","w") ## w=write ## Olmayan bir dosya adı girersem o isimde bir dosya oluşturur.
myfile.write("\nBu yazdığım metin dosyanın tüm içeriğini silecek. Sadece bu kalacak.")
myfile.close()



^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^Text Analyzer^^^^^^^^^^^^^^^^^^^^^^^^#dosyaya  her harfi alt alta yazdıran fonksiyon yaz.
filename = input("Enter a filename: ")
with open(filename) as f:
  text = f.read()
def count_char(text, char):
  count = 0
  for c in text:                       ### bu fonksiyon ve kodlarla bir metin belgesi içinde 'r' karakterinden kaç tane var onu buluruz.
    if c == char:   # if c in char: yazıp alttaki "r" yerine alfabeyi yazarsam, tüm karakterleri sayarım.
      count += 1                 
  return count
print(count_char(text, "r"))
#bu alttaki kodlarla birleştirebilirim. Her harfi alt alta yazıp yanına yüzde kaç kullanıldığını yazdırır.
for char in "abcdefghijklmnopqrstuvwxyz":
  perc = 100 * count_char(text, char) / len(text) # her karakteri 100 ile çarpıp metnin uzunluğuna böler. Bu değer perc adını verdik.
  print("{0} - {1}%".format(char, round(perc,2))) # burada 2, noktadan sonra kaç basamağın olacağını gösterir.



^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^CLASSES-OBJECTS^^^^^^^^^^^^^^^^^^ Quiz.py incele.^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
Öncelikle student adında bir dosya oluşturdum. (aynı dizin). İçine şunları ekledim
class School:
    def __init__(self,name,surname, age,no,devamli,gpsa):
        self.name=name
        self.surname=surname
        self.age=age
        self.no=no
        self.devamli=devamli
        self.gpsa=gpsa
    def __str__(self):
        return("Object: "+str(self.name)	+" "+str(self.surname)+" "+str(self.age)+" "+str(self.no)+" "+str(self.devamli)+str(self.gpsa))
    def sayhello(self):
        print("Merhaba benim adım "+self.name)   ###selfi unutma
    def on_honor_roll(self):
        if self.gpsa>=3.5:
            return True
        else:
            return False
        ^^^^^^
Daha Sonra kendi dosyamda şunları kullandım;
## student dosyasından "School" isimli classı alıyorum
from student import School  #### sadece import filename diyebliirm. Bu sefer her fonksiyon ya da her sınıfı kullanırken basına yazmakzorunda kalırım. Örn: student1= filename.School()
student1=School("Hasan ","Aslan",21,160102084,True)
print(student1)
student1.sayhello() # Merhaba benim adım Hasan.
del student1 ##öğrenciyi siler
print(student1.on_honor_roll())## bana 


^^^^^^Başka örnekler
class Employee:
    raise_amount = 1.04
    num_of_emps = 0

    def __init__(self, first, last, pay):
        Employee.num_of_emps += 1
        self.first = first
        self.last = last
        self.pay = pay
        self.email = first.lower() + "." + last.lower() + "@company.com"
        self.record_order = Employee.num_of_emps

    def fullinfo(self):
        print(
            f"Name:{self.first} \nSurname: {self.last} \nE-mail: {self.email} \nPay: {self.pay}$ \nRegister Order: {self.record_order}")

    def apply_raise(self):
        self.pay = int(self.pay * self.raise_amount)
        print(f"{self.first}'s pay raised to {self.pay}$.")

    @classmethod  # init olanlar regular methoddu.
    def set_raise_amount(cls, new_amount):  # cls standart self gibi. Amount ise benim yeni değişkenim- parameterim
        cls.raise_amount = new_amount  # zam oranını istediğim gibi değiştirebilirim ve bu tüm çalışanlar için geçerli.
        print(f"Raise amount raised %{new_amount * 100 - 100}. New value: {new_amount}")  # Employee.set_raise_amount(1.5)
    # Bunun yerine normal bir fonksiyonda self parameteri yanında amount parameteri verip herkese istediğim değerde zam yapabilirim.
    # Ama o zaman da her zam uyguladığımda tek tek parantez içinde değeri argument olarak vermem gerekir.


    @classmethod
    def add_user_from_string(cls, user_string):
        first, last, pay = user_string.split('-')
        pay = int(pay)
        return cls(first, last, pay)


str_hasan = "Hasan-Aslan-33333"
hasan = Employee.add_user_from_string(str_hasan)
hasan.fullinfo()
hasan.apply_raise()
Employee.set_raise_amount(3)
hasan.apply_raise()



print(hasan.__dict__)

^^^^Access Specifiers public,private, protected
class Person:
    def __init__(self,name):
        self._name = name #protected
        self.__name = name #private. 

hasan = Person("Hasan")
print(hasan._Person__name) #private erişmek


^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^ print(object.__dict__) classın full obje bilgisini verir.
^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^YARARLI MODÜLLER^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
from collections import OrderedDict
my_dicti = OrderedDict()
my_dicti['Hasan'] = 'first'
my_dicti['Seda'] = 'second'
for key,value in my_dicti.items():
	print(key,value)



^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^CLASS INHERITANCE^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
Başka bir dosyadan bir sınıf alıp bunu kendi sınıfım içindeki fonksiyonlara dahil etmek istiyorum. Bu yüzden 
from baskadosya import *


class Car():
    def __init__(self, make, model, year):
        self.make = make
        self.model = model
        self.year = year
        self.odometer_reading = 0


    def get_descriptive_name(self):
        long_name = str(self.year) + ' ' + self.make + ' ' + self.model
        return long_name.title()


    def read_odometer(self):
        print("This car has " + str(self.odometer_reading) + " miles on it.")


    def update_odometer(self, mileage):
        if mileage >= self.odometer_reading:
            self.odometer_reading = mileage
        else:
            print("You can't roll back an odometer!")

    def increment_odometer(self, miles):
        self.odometer_reading += miles

class Electric(Car):
    def __init__(self,make,model,year):
        super().__init__(make,model,year)

first = Electric('tesla','model s',1998)
print(first.get_descriptive_name())
first.read_odometer()
first.update_odometer(120)
first.read_odometer()




 ^^^^^^^^^^^^^^^^^^^^^^^^^^GLOBAL LOCAL VARIABLES^^^^^^^^^^^^^^^^^^^^^^^33
 ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
 ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^PATHLiB^^^^^^^^^^^^^^^^^^^
 from pathlib import Path # P büyük yani bu bir class
 mypath=Path("My_path")
 print(mypath.exist())  ## exist yerine kullanabileceklerim:
 mkdir, rmdir
 ^^^^
 mypath=Path()
 for file in mypath.glob("*"):
 	print(file)   #dizindeki dosyaları yazdırır. *.uzantı şeklinde de aranabilir.


^^^^^^^^^^^^^^^^^^^^^İLERİKİ ÇALIŞMALARIM ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^

#Bu dosyada bazı kullanışlı formüllerim var.
#En büyük sayıyı yazdırır.Listeyi değiştirmez sadece en büyüğünü görürüm.
numbers=[1,2,3,4,5] 
max=numbers[0]
for number in numbers:
     if number>max:
          max=number
print(max)
#bu kod ise numbers dizisindeki sayılar kadar alt alta "x" yazdırır.
for item in numbers:
    output=""
    for a in range(item):
        output+="x"
    print(output)
^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
mydic={"1":"one","2":"two","3":"three","4":"four"}
phone=input("Enter your numbers please: ")
output=""                                     ##yazdığım numaraları yazıyla yazar.
for number in phone:
     output+=mydic.get(number,"!")+" "
print(output)

^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
msg=input(">")
words=msg.split()
dic={
    "happy":":)","sad":":(("
}
output=""
for a in words:
    output+=dic.get(a,a)+" "
print(output)
^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^3
class Hm:
    def __init__(self,name):
        self.name=name

    def talk(self):
        print(f"Hi! I am {self.name}")
hasan=Hm("Hasan")
hasan.talk()

^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^Openpyxl modülü-excel^^Proje olarak kaydettim.^^^^^^^^^^^^^^^^^
^
import openpyxl as xl
wb= xl.load_workbook('transactions.xlsx') # dosyamı yükledim.
sheet= wb['Sheet1'] 
cell= sheet['a1']   # satır ve sütunu tanımlıyoruz.Alttaki kod da aynı görevi görür.
cell =sheet.cell(1,1) #bu kod
print(cell.value)

for row in range(1, sheet.max_row+1):
    cell=sheet.cell(row,3)          # row=satır. ilk satırdan başlar. ilk satırın 3. sütunu(row,3), sonra 2,3 3,3 olarak geçer.  
    print(cell.value)

#####
for row in range(2,sheet.max_row+1): # 2 yazmamın sebebi, ilk hücrenin int olmaması.
    cell=sheet.cell(a,3)
    corrected=cell.value*0.9 # değeri çarptım
    newcell=sheet.cell(row,4)  #row satır.
    newcell.value=corrected # çarpılan değeri yeni hücreye ekledim

wb.save('dosya adi')


^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^Lambda^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
from functools import reduce # yada functions
a = lambda x: x+2   #kısa tek seferlik fonksiyon gibi
print(a(herhangi bir sayı)) ## 

numbers = [1,2,3,4,5,6,7,8]
print(list(filter(lambda x : x%2==0, numbers)) # sadece çift sayıları yazdırır.
print(list(map(lambda x : x%2==0, numbers))) # Her bir elementin çift olup olmadğını boolean olarak verir.
nums = reduce(lambda x,y:x*y,numbers) #Bütün sayıları birbiriyle carpar

^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^Generators^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
def my_loop():
    i = 10
    while i >0:
        yield i  ## aynı değişkeni sürekli oluşturur.
        i -=1


for a in my_loop():
    print(a)

def numbers(x):
  for i in range(x):
    if i % 2 == 0:    # çift sayı bulunca onu değişken yapacak.
      yield i

print(list(numbers(11))) # [0,2,4,6,8] 

^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^Decorators^^^^^^^^^^^3
def my_decor(world):
    def inside_decor():
        print("=============")
        world()
        print("=============")
    return inside_decor

def world():
    print("HELLO WORLD!")

print_my_text = my_decor(world) #atıf
print_my_text()
#===============
#HELLO WORLD!
#===============
#Böyle bir output verir. Atıflı satırlar yerine world fonksiyonu üzerine ^@^ işareti kullanarak yazdırabilirim. Örnek aşağıda
^^^^^^^^^^^^^^^^^^^^^^^^

def decor1(func):
  def wrap():
    print("============")
    func()
    print("============")
  return wrap

@decor1
def print_text():
  print("Hello world!")
print_text()

^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^Recursion^^^^^^^^^ # fonksiyonun kendi içinde yinelenmesi.
5! = 5 * 4! # bir sayının faktöriyeli = sayının kendisinin saf hali ve bir eksiğinin faktöriyeli

def factorial(x):
    if x == 1:
        return x                      #base case olmazsa hata verir.1 sayısı bize yine 1 sayısını verir. burası olmak zorunda..
    else:
        return x*factorial(x-1)
print(factorial(5)) ##120  in range(5) gibi düşün. ilk aşama, 5*5-1=20, x = 20, 20*4-1 = 60, 60*3-1=120, 120*2-1=120

^^^^#başka bir Örnek
def fib(x):
  if x == 0 or x == 1:
    return 1
  else:
    return fib(x-1) + fib(x-2)
print(fib(4))# fibonacci 4.sayısı. Fibonacci sayılarına bak

                                             '''                  fib(4)
                                                      ¡----------------^----------------¡
                                                 fib(3)             +              fib(2)
                                               ¡------^------¡                    ¡--------^-----¡
                                            fib(2)   +  fib(1)         fib(1)   +   fib(0)
                                          ¡----^----¡           |                  |                |
                                        fib(1)+ fib(0)   |                  |                |
                                          |            |         |                  |                |
                                         1     +    1   +  1       +        1       +     1
                                         '''




^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^itertools^^^^^^^^^^^^^^^^^^^^3
from itertools import...
'''
..count # belli bir değerden itibaren saymaya başlar.
..cycle  infinitely iterates through an iterable (for instance a list or string).
.. repeat bir objeyi belirttiğim sayıda ya da sınırsız tekrarlar.
.. takewhile , filter(lambda x ) gibi
.. accumulation
..chain - combines several iterables into one long one;
accumulate - returns a running total of values in an iterable.
'''
from itertools import product, permutations
letters = ("A","B")
print(list(product(letters,range(3)))) # her bir harfi bir sayıyla yazar.
print(list(permutations(letters))) # kombinasyon


^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
import os 
clear = lambda: os.system('clear')
clear()
^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^





^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^ Magic Operators^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
class Vector2D():
    def __init__(self,x,y):
        self.x = x
        self.y = y

    def __add__(self,other):
        return Vector2D(self.x+other.x,self.y+other.y)

first = Vector2D(3,4)
second = Vector2D(4,5)
result = first+second
print(result.x,result.y)


__sub__ for -
__mul__ for *
__truediv__ for /
__floordiv__ for //
__mod__ for %
__pow__ for **
__and__ for &
__xor__ for ^
__or__ for |


^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^Regular Expressions^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
import re
pattern = r"spam"

if re.match(pattern, "spamspamspam"): #match string baştaysa eşleşir.-- re.search, re.findall
    print("Match")

str = 'My name is David. Hi David'
pattern = r'David'
newstr = re.sub(pattern,"Amy", str) ## patternı aldım. Amy'i yaptım. Str içindeki pattern. Bu sırayla yapılıyor.
print(newstr)

^^^^^Metacharacters^^^^
import re
pattern = r"gr.y"   #nokta demek, herhangi bir karakter demek.
if re.match(pattern,"grey"):
    print("Match 1")  #True
if re.match(pattern,"gray"):
    print('Match 2') #True

^^^Bir diğer metacharacterler ise '^' ve '$'. '^' : stringin başı, dolar işareti ise sonu demektir.
pattern = r'^gr.y$'
if re.match(pattern,"grey"):
    print("Match 1") # True
if re.match(pattern,"stringray"): # False

^^^^^^^Character Classes^^
#Köseli parantezle kullanılır. Parantez içindeki herhangi bir karakter, metinde var mı diye bakarız.
pattern = r"[aeıioöuü]"
if re.match(pattern,"hasan"):
    print("match 1") #True
if re.match(pattern,"rhytm blb dfs"):
    print("Match 2") ## False

^Bunun içim aynı zamanda range işlevi kullanılabilir. [a-z], [A-Z],[0-9] aralıklardaki bütün karakterlerdir.
pattern = r"[A-Z][A-Z][0-9]" # 3 karakter aranacak. İlk ikisi A-Z arası, üçüncü ise 0-9 arası olmalı.
if re.match(pattern, "LS8"): # True
if re.match(pattern,"E3") #False. Çünkü sadece 2 karakter.

^^^!!!Dikkat etmem gereken şey; yukarıda hep search değil match kullanıyorum. Tamamen eşleşme olmalı.
^^Şimdi re.search kullanacağım.
pattern =r"[^A-Z}" # ^ işareti parantezin içinde. Bu demek oluyor ki, durum tersine dönecektir.Uppercase harici stringleri arayacaktır.
if re.search(pattern,"ABCDEFGH") # False. Çünkü hepsi büyük
if re.search(pattern, "ABCDEFgh") #True.

#^ karakteri başta kullanılmadığı zaman anlamsızdır.
^^^^^More Metacharacters^^^^
#Bazı metacharacterler; + * ? { }. Bunlar tekrarı belirler.
* ='zero or more repetitions of previous thing' #previous thing can be anything in PARANTHESES.
pattern = r"^egg(spam)*"
if re.match(pattern,'egg') # True
if re.match(pattern,"eggspameggspam") # True
# * parantez dışındakinden varsa, parantez içi olmasa da olur demek. ama +, parantez içini en az bir kere bulmak ister.
pattern = r"^egg(spam)+"
if re.match(pattern,"egg") # False
if re.match(pattern,"eggspam") # True.
^ ^ işareti olduğu için egg ile başlamak zorunda.

? ='0 ya da 1 tekrar'
pattern = r"ice(-)?cream" # bir tane ya da 0 tane - işaretli durumlar için
if re.match(pattern,"ice-cream") or re.match(pattern,"icecream") # True

{} ="2 numara arasındaki tekrarı ölçer"
açıkla = '''The regex {x,y} means between x and y repetitions of something.
 Hence {0,1} is the same thing as ?.If the first number is missing, it is taken to be zero.
  If the second number is missing, it is taken to be infinity.'''

pattern = r"9{1,3}$"
if re.match(pattern,"9") # True
"999" = True
"9999"= False



^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^Groups^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^

import re

pattern = r"a(bc)(de)(f(g)h)i"

match = re.match(pattern, "abcdefghijklmnop")
if match:
   print(match.group()) ## group(0) ile aynı. Tüm stringi yazar.
   print(match.group(0))
   print(match.group(1))#bc
   print(match.group(2))#de
   print(match.group(3))#fgh
   print(match.groups())#bc, de, fgh, g


^^^^^^Gruplarda ya/ ya da kullanmak
pattern= r'^gr(a|e)y'
if re.match(pattern,"gray") or re.match(pattern,"grey") # True
if re. match(pattern, "griy")# False

^^^^^^^^^Special Sequences^^^^

pattern = r'(.+) \1' # parantez içi nokta; yani herhangi bir şeyi arıyorum. + işareti ise en az 1 demekti.(yukarıda öğrendim)
# bu parantez sonrasına backslash\ koydum ve 1-99 arası bir sayı girdim. Yani herhangi bir dizi verdiğim sayı kadar tekrar edecek mi onu ölçüyorum.
if re.match(pattern,"word word"): # True

^^^^^
Başka special sequnces:  # \d = digits, \s = whitespaces [ \t\n\r\f\v], \w = word characters : a-Z,0-9.
# Bu harfleri büyük yaparsam tam tersini ararım. Örn: \D = digits olmayan karakterler.
pattern = r"(\D+\d)"
if re.match(pattern, "Hi 999!") # True
if re.match(pattern, " ! $?")# False. Çünkü ikinci şartım bir digit bulmaktı. Sadece ilk şartım doğru.

^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^Additional Special Sequences^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
# \A \Z \b
'''
The sequences \A  \Z match the beginning and end of a string, respectively.
The sequence \b matches the empty string between \w and \W characters, or \w characters and the beginning or end of the string.
Informally, it represents the boundary between words.
The sequence \B matches the empty string anywhere else.
'''
pattern = r"\b(cat)\b"
"it is a cat" = True, "it is a <cat>" = True, "we are scattered" = False #False, çünkü \b aradığına uymuyor.
#\b whitespaces olanlara uyar. Yukarıda anlattım.

^^^^^^^^^^^^^^^^^^^^^^^^^^^^^E-mail Extraction Denemesi^^^^
str = "Please contact aslanhassan98@gmail.com for assistance."
pattern = r"([\w\.-]+)@([\w\.-]+)(\.[\w\.]+)"  #bla@bla.bla
match = re.search(pattern, str)
if match:
    print(match.group()) # Maili yazdıracak. Çoklu mail için re.findall

^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
^^^^^^^^^^^^^^Blender deneme^^^^^^^^^^^^^^^^^^
import bpy
from random import randint

bpy.ops.object.select_all(action='SELECT')
bpy.ops.object.delete(use_global=False)

for a in range(500):
    x,y,z = randint(-20,20),randint(-20,20),randint(-2,2)
    bpy.ops.mesh.primitive_monkey_add(location=(x,y,z))
    bpy.ops.object.modifier_add(type="SUBSURF")
    bpy.context.object.modifiers["Subdivision"].render_levels =3
    bpy.context.object.modifiers["Subdivision"].levels =3
    bpy.ops.object.shade_smooth()

^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
^^^^^^^^^^^^^^OS module^^^^^^^^^^
^^^Automate parsing and renaming of multiple files...
#Bir klasöre Earth-Our Solar System-1.mp4 şeklinde 6-7 dosya koydum.
import os
os.path.exist('path')
for f in os.listdir():
	print(f)### dizindeki dosyaları yazdırır.
	print(os.path.splitext(f)) # uzantıları ayrı yazdırır.
	file_name, file_ext = os.path.splitext(f)
	f_title, f_course, f_num = file_name.split('-')
	#gereksiz boşlukları kaldırmak için ise;
	f_title = f_title.strip()
	f_course = f_course.strip()
	f_num = f_num.strip().zfill(2)
	new_name = f"{f_num}-{f_title}{file_ext}"
    os.rename(f,new_name)
 	#Yeni dosya adı: 01-Earth.mp4
^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^JSON^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
import json

numbers = [2,3,4,5,6,7,8,9,10]
filename = 'numbers.json'
with open(filename,'w') as f:
	json.dump(numbers,f) #bir dosyaya kaydettim. geri yüklemek için ise:

with open(filename) as f:
	numbers = json.load(f) #geri yükledim.

^^^^^Kullanıcı adını depolayan basit bir fonksiyon
filename = 'username.json'
try:
	with open(filename) as f:
		username = json.load(f)
		print(f'Welcome, {username}!')
except:
	username = input('Enter your username please: ')
	with open(filename,'w') as f:
		json.dump(username,f)
	print('Next time we\'ll remember you!')

########################################################################3
#başka bir kullanımı da var
data = '{"firstName":"Hasan","secondName":"Aslan"}'
y = json.loads(data)
print(y['firstName'])

########################################################################3
#veritabanımda 10 kullanıcı ve bilgileri var. Users.json
with open('users.json','r') as f:
    data = json.load(f)


for user in range(len(data)):
    print(f"Full Name: {data[user]['name']}  Username: {data[user]['username']}  E-mail: {data[user]['email']}")


^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^SQLITE^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
#Sqlite3 programını ve bir örnek datayı indir
#connection.commit() kaydeder.
import sqlite3
#veritabanları klasörümde chinook.db var

connection = sqlite3.connect('chinook.db') #veritabanıma bağlanıyorum.

#veritabanında sorgu için * = tüm kolonları getir.
#customers tablosundan veri çekeceğim.
cursor = connection.execute("select * from customers")
# yada ("select FirstName, LastName from customers")

for row in cursor:
    print(f"Full Name: {row[1]} {row[2]}")

connection.close()

^^^^^^^^^^^^Tabloların adını öğrenme
def sql_fetch(con):

    cursorObj = con.cursor()

    cursorObj.execute('SELECT name from sqlite_master where type= "table"')

    print(cursorObj.fetchall())



########################################################################
^^Where ve select operatorleri
#("select * from customers where city ='Prague' or city = 'Berlin")

########################################################################
^^order operatorü
#alfabetik sıralama ->asc = ascending . tersi ise desc = descending
("select * from customers where city ='Prague' or city = 'Berlin' order by FirstName")
("select * from customers where city ='Prague' or city = 'Berlin' order by FirstName desc")

########################################################################
#şehirleri ve kaç tane oldugunu yazdıracagım.
cursor = connection.execute("select city,count(*) from customers group by city order by city") #şehre göre sırala ve alfabetik olsun.
#cursor = connection.execute("select city,count(*) from customers group by city order by count(*) desc") sayıya göre en azdan sırala.
#cursor = connection.execute("select city,count(*) from customers group by city having count(*)>1 order by count(*) desc") aynı şehir birden fazla ise onları en az sayıdan başlayarak sırala.
for row in cursor:
    print(f"{row[0]} = {str(row[1])} ")

########################################################################
^^Like operatorü
cursor = connection.execute("select FirstName, LastName from customers where FirstName like '%a%'") #isminde a harfi geçenler.
# ismi a ile başlayanlar için ise 'a%'


########################################################################
^^insert operator
#bazı veritabanlarının bazı yerleri boş bırakılmamalıdır.
#insert, veritabanına veri ekler
connection.execute("insert into customers (firstName, lastName,city,email) values('Hasan','Aslan','Manisa','aslanhassan98@gmail.com')")
#bunu bir fonksiyon yapıp, değerleri tırnak içinde vermeden de yaptırabilirdim.
connection.commit() # kaydet.   

########################################################################
^^update operator
connection.execute("update customers set city ='Çanakkale' where city='Manisa'") # set city = 'Çanakkale',firstname = 'Mehmet' where city = 'Manisa' or başka bir denklem olarak da yapabilirim

########################################################################
^delete operator
connection.execute("delete from customers where firstname='Hasan'")
#kayıt etmeyi unutma.

########################################################################
^join
#veritabanımda artistlerin ve albumlerin ayrı tabloları var ama idleri ilişik.
#sqlite3 programında execute kısmında;
select albums.Title, artists.Name from artists inner join albums on artists.ArtistId = albums.ArtistId
#play diyorum.
#bunu pythonda yapmak için
cursor  = connection.execute("select albums.Title, artists.Name from artists inner join albums on artists.ArtistId = albums.ArtistId")
#istediğim attribute'a göre cekebilirim.
cursor = connection.execute("select albums.Title, artists.Name from artists inner join albums on artists.ArtistId = albums.ArtistId where Name = 'Iron Maiden'")

^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^ITERATORS^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
#for döngüsü gibi
sehirler = ["Ankara","İstanbul","İzmir","Van"]

iteratorum = iter(sehirler)

print(next(iteratorum))
print(next(iteratorum))
print(next(iteratorum))
print(next(iteratorum))


^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^MATPLOTLIB^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
import matplotlib.pyplot as plt

input_values = [1,2,3,4,5]#bunları girmeden yaptığımda sonuçlar dengesiz gözükebiliyor.
squares = [1,4,9,16,25]
plt.plot(input_values,squares,linewidth=10)
plt.title('Square numbers',fontsize= 24)
plt.xlabel('Value',fontsize=14)
plt.ylabel('Square of Value',fontsize=14)
plt.tick_params(axis='both',labelsize=14)
plt.show()

^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
import matplotlib.pyplot as plt

x_values = list(range(1,1001))
y_values = [x**2 for x in x_values]
plt.scatter(x_values,y_values,cmap=plt.cm.Reds,edgecolors='none',s=40,c=y_values)
plt.title('Square Values',fontsize=14)
plt.xlabel('X doğrusu',fontsize=14)
plt.ylabel('Y doğrusu',fontsize=14)
plt.tick_params(axis='both',labelsize=10)
plt.axis([0,1100,0,1100000])
#plt.savefig('squares.png',bbox_inches='tight') #kaydeder.
plt.show()
^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
from random import choice
import matplotlib.pyplot as plt

class RandomWalk:
    def __init__(self, num_points=5000):
        self.num_points = num_points #initialize attribues of a walk.
        #all walks start at 0,0
        self.x_values = [0]
        self.y_values = [0]
    def fill_walk(self):
        while len(self.x_values)<self.num_points:
            #decide which direction to go and how far to go in that direction
            x_direction = choice([1,-1])
            x_distance = choice([0,1,2,3,4])
            x_step = x_direction*x_distance

            y_direction = choice([1, -1])
            y_distance = choice([0, 1, 2, 3, 4])
            y_step = y_direction * y_distance
            #reject moves that goes nowhere
            if x_step == 0 and y_step==0:
                continue
            #calculate the next x and y values
            next_x = self.x_values[-1]+x_step
            next_y = self.y_values[-1]+y_step

            self.x_values.append(next_x)
            self.y_values.append(next_y)

rw=RandomWalk()
rw.fill_walk()
point_numbers = list(range(rw.num_points))
plt.scatter(rw.x_values,rw.y_values,s=15,edgecolors='none',c=point_numbers,cmap=plt.cm.Blues)
plt.show()
	
^^^^^^^^^^^^^^^^^^^^^^^^^^^^pyc PYcompile^^^^^^^^^^
împort py_commpile
py_compile.compile('script.py')








^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
PyQt5 donustur
pyuic5 untitled.ui -o Done.py







^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
^^^^^^^^^^^^^^^^^^^^^^^^^^^^MACHINE LEARNING^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
from scipy import stats
import numpy
speed = [99,86,87,88,111,86,103,87,94,78,77,85,86]
x_average = numpy.mean(speed) #Mean bize ortalama değeri verir.
x_median = numpy.x_median(speed)#Bütün değerler sort edilince en ortada kalan değer mediandır.
x = stats.mode(speed) # Mode: en çok tekrar edilen değer.

^^^Standard Deviation (Standart Sapma)
 #Sigma: σ işareti	
#If values are close to each otheri standard deviation is LOW.Else, logically, it is high. speed = [24,45,212,542] Standart Deviation = 207
speed = [86,87,88,86,87,85,86] #Standard deviation = 0.9
x = numpy.std(speed) # standart sapmayı hesaplar.

^^^Variance
#Variance is another method that indicates how spread out the values are.
#In fact, if you take the square root  of the variance, it will give you the standart deviation.
#To calculate variance; 1) Find the mean(ortalama). 
#For each number, find the difference between the number and the mean. For each, take their square. 
#And finally, divide  the number len(array)-1.
 #Sigma: σ2

 ^^^^^^Percentiles
 #Grafikteki konum.Bir değerin kaçıncı yüzdede yer aldığını bulmak ya da
 #Verdiğin yüzde değerrinde hangi değerin olduğunu bulmak
 import numpy
ages = [5,31,43,48,50,41,7,11,15,39,80,82,32,2,8,6,25,36,27,61,31]
x = numpy.percentile(ages, 90)
print(x)


^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^WEB SCRAPING^^^^^^^^^^^^^^^^^^^^^^^^^^
#myarchive icinde convert_money ve scripts icide github projelerimi incele.

import requests
import json

result = requests.get("https://jsonplaceholder.typicode.com/todos")

result  = json.loads(result.text)

for a in result:
    print(a["userId"],a["id"],a["title"],a["completed"])
















