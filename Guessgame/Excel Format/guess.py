import datetime
import random
import openpyxl as xl

memory = []
user = input("Hello. I am a basic gambler. And this is a guess game. What is your name?\n")
print("Nice to meet you " + str(user))
is_correct = False
a = 1
while a > 0:
    try:
        num1 = int(input("Tell me the minimum number you want.\n"))
        num2 = int(input("and the highest one.\n"))
        made = int(input("And how many chances you ask\n"))
        break
    ##############quit opsiyonu eksik.
    except:
        print("Enter a number please, not a string!")
if num1 < num2:
    numran = random.randint(num1, num2)
else:
    numran = random.randint(num2, num1)
guessesmade = 0

input("Okay. You have " + str(made) + " chances to guess the number between " + str(num1) + " and " + str(
    num2) + "\n Hit enter if it is clear ")
while guessesmade < made:
    try:
        guess = int(input("Guess it.\n"))
        guessesmade += 1
        memory.append(guess)
        if guess < numran:
            print("No, it is higher.")
        if guess > numran:
            print("No. It is lower.")
        if guess == numran:
            break
    except:
        print("Enter a number please. ")
if guess == numran:
    is_correct = True
    print("Congratulations.You find my number trying only " + str(guessesmade) + " times!")
else:
    print("No " + str(user) + "! The number I think was " + str(numran))
    is_correct = False
if num1 == num2:
    print("But you did a cheat which I winked at :)")
wb = xl.load_workbook('guessrecords.xlsx')
sheet = wb['Sheet1']

aaa = 2
while True:
    cell = sheet.cell(aaa, 1)
    if cell.value != None:
        aaa += 1
        continue
    else:
        break
cell = sheet.cell(aaa, 1)
cell.value = user
cell = sheet.cell(aaa, 2)
cell.value = (f"{num1}-{num2}")
cell = sheet.cell(aaa, 3)
cell.value = made
cell = sheet.cell(aaa, 4)
cell.value = guessesmade
cell = sheet.cell(aaa, 5)
cell.value = str(is_correct)
cell = sheet.cell(aaa, 6)
cell.value = numran
cell = sheet.cell(aaa, 7)
cell.value = str(memory)
cell = sheet.cell(aaa, 8)
cell.value = str(datetime.datetime.now())

wb.save('guessrecords.xlsx')
input('Leaving...')