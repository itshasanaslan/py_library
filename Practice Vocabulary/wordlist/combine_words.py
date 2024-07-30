import openpyxl as xl
from pathlib import Path
from colorama import Fore,Style,init

init()
files = []
print('\n\nPlease wait.')
mypath=Path()
for file in mypath.glob("*.xlsx"):
    files.append(file)



words = {}
translated = {}

for file in files:
	try:
		wb = xl.load_workbook(file)
		sheet = wb['Sayfa1']
		for a in range(2,sheet.max_row-1):
			cell = sheet.cell(a,1)
			word = cell.value

			cell = sheet.cell(a,2)
			definition = cell.value
			
			cell = sheet.cell(a,3)
			trans = cell.value
			if word == None:
				continue
			words.update({word:definition})
			translated.update({word:trans})

	except Exception as got_error:
		input(got_error)
number =1
for item,key in words.items():
	print(f"{number}-{item} : {key}\nTR: {translated.get(item)}\n-------------------------------------")
	number+=1

my_col =1

try:
	wb = xl.load_workbook('All Words.xlsx')
	sheet = wb['Sayfa1']
	for word in words:
		my_col+=1
		cell = sheet.cell(my_col,1)
		cell.value = word

		cell = sheet.cell(my_col,2)
		cell.value = words.get(word)

		cell = sheet.cell(my_col,3)
		cell.value = translated.get(word)

except Exception as error2:
	print(error2)
wb.save('All Words.xlsx')

print(f'{Style.BRIGHT+Fore.GREEN}\n\nTüm kelimeler birleştirildi.')
input()