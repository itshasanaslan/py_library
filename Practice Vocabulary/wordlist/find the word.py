print('\n\nPlease wait.')

import openpyxl as xl
from pathlib import Path
from colorama import Fore,Style,init
from os import system


system('color 70')
system('title Find my word')
clear = lambda:system('cls')

init()
files = []

mypath=Path()
for file in mypath.glob("*.xlsx"):
    files.append(file)




words = {}
translated = {}
is_first = True
is_found=False
while True:
	clear()
	print('This file helps me to find out in which xl file my word is.\n')
	if not is_first:
		if not is_found:
			input('\nThe word you seek does not exist in your database.')


	find_my_word = input('\nType "q" to exit.\nWhich word you want me to find?\n>>>')
	if find_my_word=='q':
		exit()
	for file in files:
		filename = str(file)
		if filename =='All Words.xlsx':
			continue
		try:
			wb = xl.load_workbook(file)
			sheet = wb['Sayfa1']
			for a in range(2,sheet.max_row-1):
				cell = sheet.cell(a,1)
				word = cell.value
				word = str(word)
				is_first=False
				if word ==find_my_word:
					is_found=True
					input(f"\n\nFound your word at {file}")
					break
				cell = sheet.cell(a,2)
				definition = cell.value
				
				cell = sheet.cell(a,3)
				trans = cell.value
				if word != None:
					words.update({word:definition})
					translated.update({word:trans})

		except Exception as got_error:
			input(got_error)


input()