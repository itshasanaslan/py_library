import random
import os
from colorama import Fore, Style, init
import openpyxl as xl
import datetime
import openpyxl as xl
import time
from pathlib import Path
from sys import stdout
import win32com.client as wincl
pronounce = wincl.Dispatch("SAPI.SpVoice")



init()

red = Style.BRIGHT + Fore.RED
yellow = Style.BRIGHT + Fore.YELLOW
blue = Style.BRIGHT + Fore.BLUE
default = Style.RESET_ALL
magenta = Style.BRIGHT + Fore.MAGENTA
green = Style.BRIGHT + Fore.GREEN

arr_words = []
used_words = []
failed_words = []
wants_word_list = False
xl_files = []  # holds all xl files in a directory.

# for hangman
called_on_hangman = []  # holds the word called for hangman.
blanks = []  # for each letter in the word, creates a blank for hangman
letters = []  # will pick a random letter for hangman
this_letter_choosen = []  # if letter is choosen before, it will select a new letter.

os.system('title Aslan Practice Tool')
main_dir = os.getcwd()
os.chdir('wordlist')  # access to the word archive folder.
mypath = Path()
instruction = f"Given the definitions, you need to write the correct word.\nYou can type {red}'help'{green}.\nYou can type {red}'tr'{green} to see translation,\nYou can type {red}'h'{green} to see the word as in the hangman.\nYou can type {red}'show'{green} to see the all words.\n{red}'voice'\t{green} to enable/disable pronunciation.(Default:enabled)\nEach words gives you 5 points.Each help decreases 1 points.\nYou can press {red}fn+F11{green} for fullscreen.\nYou can press {red}Ctrl+C{green} to quit immediately."
wants_voice = True # for pronunciation.
clear = lambda: os.system('cls')

commands = f'''{yellow}
        You can try these commands:
        {red}'h'\t{yellow} to see the words.
        {red}'tr'\t{yellow} to see the translation.
        {red}'show'\t{yellow} to see all the words left.
        {red}'voice'\t{yellow} to enable/disable pronunciation.
        {red}'q'\t{yellow} to quit.
        {default}
'''


def info():
    print(f'''{green}
        in the folder named 'wordlist', there are excel files. Each of them is a word pool.
        These files contains words' info.
        Do not move and edit them.
        If you want to create a new word pool, just copy the file 'taslak' and rename the new file.
        Add your words and their info in the relevant columns.
        To combine all words in the 'All Words' pool, run the 'combine words' app and wait.
        Finally, do not move any file.
        {default}''')


while True:
    info()
    file_index = 1  # will use it when listing files.
    for file in mypath.glob("*.xlsx"):
        print(f"\t\t{red}{file_index}-{default}{file}")
        xl_files.append(str(file))
        file_index += 1

    dir = input('\n>Enter a file number: ')
    if not dir.isnumeric():
        print(f'\n{red}You need to enter a number!{default}')
        input('Press enter.')
        clear()
        continue
    dir = int(dir)
    try:
        dir = xl_files[dir - 1]
        kelime_wb = xl.load_workbook(dir)
        kelime_sheet = kelime_wb['Sayfa1']
        break
    except Exception as a:
        input()
        print(a)
        continue

os.system(f'title Aslan Practice Tool - Word Pool: {dir[:-5]}')

clear()
print(f"\n\n\n{green}")
for a in instruction:
    print(a, end='')
    stdout.flush()
    #time.sleep(0.03)
print(default)
input('\nPress enter.')
clear()

words = {}
translated = {}

for a in range(2, kelime_sheet.max_row - 1):
    cell = kelime_sheet.cell(a, 1)
    word = cell.value

    cell = kelime_sheet.cell(a, 2)
    definition = cell.value

    cell = kelime_sheet.cell(a, 3)
    trans = cell.value
    if not word == None:
        words.update({word: definition})
    if not word == None:
        translated.update({word: trans})

for a in words:
    arr_words.append(a)

new_arr_words = arr_words  # a list to hold all items permanently.
len_words = len(arr_words)  # this will be useful for the final calculation.


def produce():
    global selected_word
    selected_word = random.choice(new_arr_words)
    if selected_word in used_words:
        if len(used_words) == len(new_arr_words):
            selected_word = random.choice(failed_words)
            print(green, '\n\nFailed words are incoming.', default)
        else:
            print('Loop üstü.')
            while selected_word in used_words:
                selected_word = random.choice(new_arr_words)
                print('While loopunda seçilen kelime', selected_word)
                if selected_word not in used_words:
                    break
    used_words.append(selected_word)
    clear()

def typewriter_animate(text):
    for a in text:
        print(a,end='')
        stdout.flush()
        time.sleep(0.03)




def show_words_left():
    number = 1
    new_list = arr_words
    new_list.sort()
    print(f"{yellow}\n\t\t\t\t\t\tWORDS LEFT\n{default}", end='')
    for a in new_list:
        if number == 1:
            color = blue
            number += 1
        elif number == 2:
            color = green
            number = 1
        if a == new_list[-1]:
            print(color, a, default)
            print(default)
        else:
            print(color, a, end=' ')


def test():
    global selected_word, got_help, score, wants_word_list, round_count,wants_voice
    score = 0
    round_count = 1
    got_help = False
    selected_word = random.choice(arr_words)
    used_words.append(selected_word)
    while True:
        print(
            f'{blue}Words left:{yellow}{len(arr_words)}\t{blue}Score: {yellow}{score}\t{blue}Round: {yellow}{round_count}{default}')
        if wants_word_list:
            show_words_left()
            print('\n')
        if got_help:
            print('\nTR:', yellow, translated.get(selected_word), default, '\n')
        print(f'{red}\nDefinition:{default}', end=' ')
        print(words.get(selected_word))

        guess = input(f'{green}\nEnter the word: {default}').lower()
        if guess == selected_word:
            score += 5
            if wants_word_list:
                score -= 1
            print('\n')
            print(blue, '---------------------------------------', magenta)
            if wants_voice:
                typewriter_animate('Correct guess! Score+5. Pronuncing...')
            else:

                typewriter_animate('Correct guess! Score+5')
            
            print(blue, '\n---------------------------------------', default)
            if not got_help:
                print('TR:', yellow, translated.get(selected_word), default)
            if not wants_voice:
                input(f'{magenta}\n\npress enter.{default}')
            if wants_voice:
                pronounce.Speak(selected_word)
                print(f"{green} \ntype 'r' to listen again.{default}")
                while True:
                    inside_voice_loop = input('>')
                    if inside_voice_loop=='r' or inside_voice_loop=='R':
                        pronounce.Speak(selected_word)
                        continue
                    else:
                        break
            arr_words.remove(selected_word)
            round_count += 1
            if len(arr_words) == 0:
                input('\n\nCongratulations.You have finished the game.')
                finish()
            produce()
            got_help = False
            continue
        elif guess == 'tr':
            clear()
            if got_help:
                continue
            score -= 1
            got_help = True
        elif guess == 'h':
            clear()
            hangman(selected_word)
        elif guess=='voice':
        	clear()
        	if wants_voice:
        		wants_voice=False
        		print(f"{magenta}Pronunciation disabled.{default}")
        	else:
        		wants_voice=True
        		print(f"{magenta}Pronunciation enabled.{default}")
        elif guess == 'show':
            clear()
            wants_word_list = True
        elif guess == 'q':
            print(f"Goodbye.")
            finish()
            exit()
        elif guess == 'help':
            clear()
            print(f'''
                    type {yellow}'h'{default} for hangman clues.       {red}-1 score{default}.
                    type {yellow}'tr'{default} for Turkish translation.{red}-1 score{default}.
                    type {yellow}'show'{default}to see all the remaining words.
                ''')
            input('Press any key to continue.')
        else:
            print(f'{red}\n\tFailed. New word is coming.{default}', end='')
            score -= 1
            if selected_word in called_on_hangman:
                called_on_hangman.remove(selected_word)
            if not got_help and not wants_word_list:
                print(commands)
            used_words.remove(selected_word)
            round_count += 1
            blabla = input('Press enter.')
            if blabla=='q':
            	finish()
            elif blabla=='h' or blabla=='tr' or blabla=='show':
            	input('Not now. You need to write it on the "Enter the word" section.')
            got_help = False
            produce()
            failed_words.append(selected_word)
            clear()



def hangman(word):
    global score,blanks,letters,this_letter_choosen
    while True:
        check_blank_word = ''  # convert blanks array to str
        for a in blanks:
            check_blank_word += a
        if check_blank_word == word:
            print(red, 'You already found all words. -2 score.', default)
            score -= 2
            break
        if not word in called_on_hangman:
            blanks,letters,this_letter_choosen = [],[],[]
            for a in word:
                blanks+='_'
                letters.append(a)
            called_on_hangman.append(word)
        while True:
            letter = random.choice(letters)
            if letter not in this_letter_choosen:
                this_letter_choosen.append(letter)
                break
            else:
                continue
        for i in range(len(word)):
            if word[i] == letter:
                blanks[i] = letter
        print(red, *blanks, default)
        score -= 1
        break




def format_time():
    global output
    x = datetime.datetime.now()
    output = ''
    convert_str = x.strftime('%x')
    output += convert_str[3:5] + '/'
    output += convert_str[0:2] + '/'
    output += convert_str[6:8]
    clock_time = str(x.strftime('%X'))
    output += ' | ' + clock_time


def finish():
    os.chdir(main_dir)
    max_score = len_words * 5
    birlik = max_score / 100
    final_score = score / birlik
    print(f"\n\n{red}You have finished the game with {yellow}%{final_score} {red}success.")
    try:
        wb = xl.load_workbook('records.xlsx')
        sheet = wb['Sayfa1']
        my_row = 2

        while True:
            cell = sheet.cell(my_row, 1)
            if cell.value != None:  # searching for empty row.
                my_row += 1
                continue
            else:
                break

        raw_time = datetime.datetime.now()
        time = str(raw_time)

        cell = sheet.cell(my_row, 1)
        cell.value = dir[:-5]
        cell = sheet.cell(my_row, 2)
        cell.value = final_score
        cell = sheet.cell(my_row, 3)
        format_time()
        cell.value = output
        wb.save('records.xlsx')
        print('Data saved.')
    except Exception as err:
        print(err)

    input('Press enter.')
    exit()


test()
