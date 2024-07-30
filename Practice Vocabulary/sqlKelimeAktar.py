import openpyxl as xl
import sqlite3

words = {} #kelimeleri ve definitionu tutar
translated = {}

#excelden kelimeleri aldım
kelime_wb = xl.load_workbook("All Words.xlsx")
kelime_sheet = kelime_wb['Sayfa1']


for a in range(2, kelime_sheet.max_row - 1):
    cell = kelime_sheet.cell(a, 1)
    word = cell.value

    cell = kelime_sheet.cell(a, 2)
    definition = cell.value

    cell = kelime_sheet.cell(a, 3)
    trans = cell.value
    if not word == None:
        words.update({word: definition})
    if not word == None:
        translated.update({word: trans})


connection = sqlite3.connect("wordList.db")

for kelime in words:
	meaning = words.get(kelime)
	translation = translated.get(kelime)
	connection.execute("insert into WordList (Word, Meaning, Translation) values(?,?,?)",(kelime,meaning,translation))


connection.commit() # kaydet.   
connection.close()
input("Done")