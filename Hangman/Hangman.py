#del fonksiyonu tek seferlik
import random
import datetime
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from colorama import Style,Fore,init
from time import sleep

init()
green = Style.BRIGHT+Fore.GREEN
blue =  Style.BRIGHT+Fore.BLUE
red = Style.BRIGHT+Fore.RED
magenta = Style.BRIGHT+Fore.MAGENTA
default = Style.RESET_ALL
is_found = False
words = ["lion", "cat", "tiger", "turtle", "fox", "ant", "giraffe", "fly", "mosquito", "dog", "pigeon", "parrot",
         "elephant", "zebra", "camel", "snake","mouse"]

try:
    deneme = []
    with open('words', 'r') as f:
        deneme = f.read().split('-')
        for a in deneme:
            if a not in words:
                words.append(a)
except:
    with open("words", "w+") as f:
        words += f.read().split('-')
        print('New document created')
for a in words:
    if words.count(a) > 1:
        words.remove(a)
    if a=="" or a==" ":
        words.remove(a)

guess_word = []
secret_word = random.choice(words)
len_secret_word = len(secret_word)
alphabetcheck = "a b c ç d e f g ğ h ı i j k l m n o ö p r s ş t u ü v y z w q x"
letter_storage = []
def welcome():
    while True:
        global user
        print(magenta)
        user = input(f"Welcome. What's your name?{default} \n").strip()
        if user.lower() == "admin":
            pas_check = input("Then authenticate..........\n ")
            print('Checking.....')
            for a in range(5):
                print(a * "******" + " ")
                sleep(0.4)
            if pas_check == "8448209aA":

                print(blue+
                    "################################################################################################################")
                print(
                    "########################################ACCESS GRANTED!#########################################################")
                print(
                    "################################################################################################################")
                return admin_panel()
            else:
                print(red+
                    "################################################################################################################")
                print(
                    "#     ACCESS DENIED!     ##     ACCESS DENIED!     ##     ACCESS DENIED!     ##     ACCESS DENIED!     ##     ")
                print(
                    "################################################################################################################")
                continue
        if user.lower() == "" or user.lower() == " ":
            print(red+"Sorry mate, but I need a name!")
        elif user.lower() == "quit":
            print(blue+"See you later..")
            quit(1)
        else:
            print(magenta+"Nice to meet you ", user)
            i_play = input("Type or press anything to play. Or 'help' to get help.").strip()
            if i_play.lower() == "help":
                print('''                This is a hangman game coded by Hasan Aslan. 
                We choose a random word and you try to guess. You can write only one letter
                each time. You have only 9 wrong chances. If you write the letter you already
                write, it won't count.''')
                print(user, ", the game is starting by the way... :) just hit the enter.")
                input()
                break
            else:
                print("Let's play ", user)
                break


def admin_panel():
    print(red+'''Welcome Sir! These are commands you can apply;
                 Type see to see all words
                 Type add to add words
                 Type del to del words
                 Type reveal to see the hidden word
                 Type new to create a new secret word
                 Type 1 to create and see the new word
                 Type quit to leave and continue''')
    while True:
        global secret_word
        add_words = ""
        admin_input = input(green+"User/Admin>>>>>>").lower()
        if admin_input == "see":
            print(f"Here is the words: {list(words)}")
            continue
        if admin_input=="reveal":
            print(f"The hidden word is {secret_word}")
            continue
        if admin_input=="new":
            secret_word=random.choice(words)
            print("Secret word has changed")
            continue
        if admin_input=="1":
            secret_word = random.choice(words)
            print(f"New hidden word is {secret_word}")
        if admin_input == "add":
            print("Type \'done\' to stop.")
            return func_to_add_words()
        if admin_input == "del":
            return func_to_del_words()
        if admin_input == "quit":
            try:
                with open("words","r+") as f:
                    text=f.read()
                    for item in words:
                        if item not in text:
                            f.write(str(item) + "-")
            except:
                print("Couldn't handle")
                quit()


        break

def func_to_del_words():
    print(red+"               <Word deleting section> ")
    print(green+"Just write the words with order... lion cat mouse")
    while True:
        del_which = input("Enter the words:  ")
        if del_which == "done":
            return admin_panel()
        try:
            del_which = del_which.split()
            for a in del_which:
                words.remove(a)
                print(f"{a} has been removed")
        except ValueError:
            for b in del_which:
                print(b,"is not in the list..")
            continue
        except:
            print("There must be some kind of error sir. I can't do that!")
            continue


def func_to_add_words():
    print(green+"             <Word adding section> ")
    while True:
        add_words = input(blue+"write the words>>>>> ")
        if add_words == "done":
            print("Completed...")
            return admin_panel()
        else:
            add_words = add_words.split()
        try:
            for a in add_words:
                if a not in words:
                    words.append(a)
                    print(f"{a} added.")
                else:
                    print(+f"The word {a} is already exists sir!")
                    continue
        except:
            print("Sir, I can't add words. There must be some kind of error")
            continue


def blanks():
    for a in secret_word:
        guess_word.append('_')
    print(magenta+"The word you need to guess has ", len_secret_word, " characters.")
    print("You can only use a single character at once.")
    print(*guess_word, sep=" ")


def ingame():
    global score
    global is_found
    score = 1
    while True:
        guess = input(magenta+"Guess a letter: ").lower()
        if guess == "admin":
            guess = input("Then authenticate>  ")
            for a in range(5):
                print(a * "******" + " ")
                sleep(0.4)
            if guess == "8448209aA":
                print(blue +
                      "################################################################################################################")
                print(
                    "########################################ACCESS GRANTED!#########################################################")
                print(
                    "################################################################################################################")
                return admin_panel()
            else:
                print(red +"################################################################################################################")
                print(
                    "#     ACCESS DENIED!     ##     ACCESS DENIED!     ##     ACCESS DENIED!     ##     ACCESS DENIED!     ##     ")
                print(
                    "################################################################################################################")
                continue

        if guess.isnumeric():
            print(red+"Numbers are not allowed ", user)
        elif guess == "quit":
            print(blue+"Called a quit code. Leaving..")
            quit()
        elif guess == secret_word:
            is_found = True
            print(green+"Perfect You found it.")
            break
        elif len(guess) != 1:
            print(green+"Use only a single letter at once. Use these; ", alphabetcheck)
            score += 1
        elif guess in letter_storage:
            print(green+"You already tried it.")
        else:
            letter_storage.append(guess)
            if guess in secret_word:
                print(green+"You guessed correctly. Keep going on", user + "!")
                for i in range(0, len_secret_word):
                    if secret_word[i] == guess:
                        guess_word[i] = guess
                print(*guess_word, sep=" ")
                if not '_' in guess_word:
                    print(green+"You won")
                    is_found = True
                    break
            else:
                print(red+"The letter is not in the word.Try again!",green+"\n You have ", 9 - score, " guesses left")
                print(*guess_word, sep=" ")
                score += 1
            if score > 9:
                print(red+"Sorry, you are out of guesses.The secret word was: ", secret_word)
                break


welcome()
blanks()
ingame()

a = datetime.datetime.now()
time = str(a)


def file_caller():
    while True:
        global sheet, wb, is_created
        is_created = False
        try:
            wb = xl.load_workbook('records.xlsx')
            sheet = wb['Sheet1']
            break
        except:
            wb = Workbook()
            worksheet = wb.active
            worksheet.title = 'Sheet1'
            wb.save('records.xlsx')
            wb = xl.load_workbook('records.xlsx')
            sheet = wb['Sheet1']
            sheet.row_dimensions[1].height = 30
            sheet.column_dimensions['A'].width = 12
            sheet.column_dimensions['B'].width = 9
            sheet.column_dimensions['C'].width = 18
            sheet.column_dimensions['D'].width = 25
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 20
            sheet.column_dimensions['G'].width = 20
            sheet.column_dimensions['H'].width = 20

            is_created = True
            for a in range(1, 100):
                cell = sheet.cell(a, 1)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
                cell = sheet.cell(a, 2)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
                cell = sheet.cell(a, 3)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
                cell = sheet.cell(a, 4)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
                cell = sheet.cell(a, 5)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
                cell = sheet.cell(a, 6)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
                cell = sheet.cell(a, 7)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
                cell = sheet.cell(a, 8)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
                cell = sheet.cell(1, a)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)

            break
        finally:
            sheet = wb['Sheet1']


def my_games_records():
    if is_created:
        cell = sheet.cell(1, 1)
        cell.value = "User"
        cell.font = Font(color=colors.BLUE, italic=False, bold=True)
        cell = sheet.cell(1, 2)
        cell.value = 'Guess Correct'
        cell.font = Font(color=colors.BLUE, italic=False, bold=True)
        cell = sheet.cell(1, 3)
        cell.value = 'Chances Used'
        cell.font = Font(color=colors.BLUE, italic=False, bold=True)
        cell = sheet.cell(1, 4)
        cell.value = 'Letters Used'
        cell.font = Font(color=colors.BLUE, italic=False, bold=True)
        cell = sheet.cell(1, 5)
        cell.value = 'Secret Word'
        cell.font = Font(color=colors.BLUE, italic=False, bold=True)
        cell = sheet.cell(1, 6)
        cell.value = 'Time'
        cell.font = Font(color=colors.BLUE, italic=False, bold=True)
        wb.save('records.xlsx')

    else:
        print('Saved!')


file_caller()
my_games_records()

aaa = 2
while True:
    cell = sheet.cell(aaa, 1)
    if cell.value != None:
        aaa += 1
        continue
    else:
        break

cell = sheet.cell(aaa, 1)
cell.value = user
sheet.row_dimensions[aaa].height = 30
cell = sheet.cell(aaa, 2)
cell.value = str(is_found)
cell = sheet.cell(aaa, 3)
cell.value = score
cell = sheet.cell(aaa, 4)
cell.value = str(letter_storage)
cell = sheet.cell(aaa, 5)
cell.value = secret_word
cell = sheet.cell(aaa, 6)
cell.value = (time[:19])
print(">>>>>>>>>>>>>>>>>>>>Game over!<<<<<<<<<<<<<<<<<")
wb.save('records.xlsx')







